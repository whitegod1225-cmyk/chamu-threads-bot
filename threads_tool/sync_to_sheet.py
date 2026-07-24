"""
sync_to_sheet.py
Threads投稿のメトリクスをxlsxのPDCAダッシュボードに書き込む。

前提:
  - threads_metrics_cache.json が同フォルダにある（/fetcherが更新する）
  - post_meta.json が同フォルダにある（投稿時に手動追記する）
  - content_pdca_base.xlsx が同フォルダにある（init_xlsx.pyで作成する）
  - openpyxl がインストール済み: pip install openpyxl

使い方:
  python sync_to_sheet.py
  → threads_metrics_cache.json + post_meta.json をマージして xlsx に書き込む

  python sync_to_sheet.py --from-meta
  → skills/queue/post-history-meta.md を読んで xlsx に書き込む（手動メタ不要）

  python sync_to_sheet.py --create-if-missing
  → xlsx が存在しなければ自動作成してから書き込む
"""

import os
import sys
import json
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl がインストールされていません。\npip install openpyxl")

HERE = os.path.dirname(__file__)
CACHE_PATH = os.path.join(HERE, "threads_metrics_cache.json")
META_PATH = os.path.join(HERE, "post_meta.json")
XLSX_PATH = os.path.join(HERE, "content_pdca_base.xlsx")
POST_HISTORY_META = os.path.join(HERE, "..", "skills", "queue", "post-history-meta.md")

HEADER_FILL = PatternFill("solid", fgColor="1C2333")
HEADER_FONT = Font(bold=True, color="EDE7D8", size=10)
ALT_FILL = PatternFill("solid", fgColor="F0EBE0")
BORDER_COLOR = "CDBE9E"
thin = Side(style="thin", color=BORDER_COLOR)
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    ("投稿ID", 10),
    ("日付", 12),
    ("テーマ", 28),
    ("型", 8),
    ("パターン", 10),
    ("大分類", 14),
    ("Views", 10),
    ("Likes", 8),
    ("Comments", 10),
    ("Reposts", 9),
    ("Quotes", 8),
    ("Resonanceスコア", 14),
    ("質スコア(%)", 11),
    ("permalink", 44),
]


def load_json(path, default):
    if not os.path.exists(path):
        return default
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def create_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDCAダッシュボード"

    ws.row_dimensions[1].height = 20
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"新しいxlsxを作成しました: {path}")
    return wb


def build_rows_from_meta():
    """post-history-meta.md のMarkdownテーブルを読んで行データを返す"""
    import re
    path = os.path.abspath(POST_HISTORY_META)
    if not os.path.exists(path):
        sys.exit(f"post-history-meta.md が見つかりません: {path}\n先に make_meta.py を実行してください。")

    with open(path, encoding="utf-8") as f:
        text = f.read()

    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|") or line.startswith("| 番号") or line.startswith("|---"):
            continue
        cols = [c.strip() for c in line.strip("|").split("|")]
        if len(cols) < 14:
            continue
        (post_num, posted_at, theme, category,
         views, likes, replies, reposts,
         score, quality, hook_type, cta_type, fetched, post_id) = cols[:14]

        def safe_int(v):
            try: return int(v)
            except: return 0

        def safe_float(v):
            try: return float(v)
            except: return 0.0

        v = safe_int(views)
        l = safe_int(likes)
        r = safe_int(replies)
        rp = safe_int(reposts)
        sc = safe_int(score) if score != "-" else (l * 3 + r * 5 + rp * 2)
        ql = safe_float(quality) if quality != "-" else round(sc / v * 100, 2) if v > 0 else 0.0

        # カテゴリから型・パターン・大分類を推定
        cat_m = re.match(r'([A-Z]+)[（(](.+?)[）)]', category)
        cat_type = cat_m.group(1) if cat_m else category[:2]
        cat_detail = cat_m.group(2) if cat_m else category

        date_str = posted_at[:10] if posted_at != "-" else ""
        permalink = f"https://www.threads.net/post/{post_id}" if post_id not in ("-", "") else ""

        rows.append({
            "投稿ID": post_id if post_id != "-" else post_num,
            "日付": date_str,
            "テーマ": theme,
            "型": cat_type,
            "パターン": hook_type if hook_type != "-" else "",
            "大分類": cat_detail,
            "Views": v,
            "Likes": l,
            "Comments": r,
            "Reposts": rp,
            "Quotes": 0,
            "Resonanceスコア": sc,
            "質スコア(%)": ql,
            "permalink": permalink,
        })

    rows.sort(key=lambda r: r.get("日付") or "", reverse=True)
    print(f"[from-meta] {len(rows)}行を読み込みました")
    return rows


def build_rows():
    cache = load_json(CACHE_PATH, {})
    meta_list = load_json(META_PATH, [])
    meta_by_id = {
        r.get("投稿ID"): r for r in meta_list
        if isinstance(r, dict) and "投稿ID" in r and "_comment" not in r
    }
    meta_by_permalink = {
        r.get("permalink"): r for r in meta_list
        if isinstance(r, dict) and r.get("permalink") and "_comment" not in r
    }

    rows = []
    for post_id, metrics in cache.items():
        if not isinstance(metrics, dict):
            continue
        permalink = metrics.get("permalink") or metrics.get("id") or ""
        meta = meta_by_id.get(post_id) or meta_by_permalink.get(permalink) or {}

        views = metrics.get("views") or 0
        likes = metrics.get("likes") or 0
        replies = metrics.get("replies") or 0
        reposts = metrics.get("reposts") or 0
        quotes = metrics.get("quotes") or 0
        resonance = likes * 3 + replies * 5 + reposts * 2
        quality = (resonance / views * 100) if views > 0 else 0.0

        timestamp = metrics.get("timestamp") or metrics.get("created_at") or ""
        date_str = ""
        if timestamp:
            try:
                date_str = datetime.fromisoformat(timestamp.replace("Z", "+00:00")).strftime("%Y-%m-%d")
            except Exception:
                date_str = str(timestamp)[:10]

        rows.append({
            "投稿ID": meta.get("投稿ID") or post_id,
            "日付": date_str,
            "テーマ": meta.get("テーマ") or "",
            "型": meta.get("型番号") or "",
            "パターン": meta.get("パターン") or "",
            "大分類": meta.get("悩み大分類または販売ジャンル") or "",
            "Views": views,
            "Likes": likes,
            "Comments": replies,
            "Reposts": reposts,
            "Quotes": quotes,
            "Resonanceスコア": resonance,
            "質スコア(%)": round(quality, 2),
            "permalink": permalink,
        })

    rows.sort(key=lambda r: r.get("日付") or "", reverse=True)
    return rows


def write_rows(ws, rows):
    max_existing = ws.max_row
    existing_permalinks = set()
    for row in ws.iter_rows(min_row=2, max_row=max_existing, values_only=True):
        if row and row[13]:
            existing_permalinks.add(str(row[13]))

    written = 0
    for data in rows:
        permalink_val = str(data.get("permalink") or "")
        if permalink_val in existing_permalinks and permalink_val:
            continue

        next_row = ws.max_row + 1
        col_keys = [c[0] for c in COLUMNS]
        is_alt = (next_row % 2 == 0)

        for col_idx, col_name in enumerate(col_keys, 1):
            val = data.get(col_name, "")
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if is_alt:
                cell.fill = ALT_FILL
        written += 1

    return written


def main():
    parser = argparse.ArgumentParser(description="Threadsメトリクスをxlsxに同期")
    parser.add_argument("--create-if-missing", action="store_true",
                         help="xlsxが存在しなければ自動作成する")
    parser.add_argument("--from-meta", action="store_true",
                         help="post-history-meta.mdを読んでxlsxに書き込む（手動メタ不要）")
    args = parser.parse_args()

    if not args.from_meta and not os.path.exists(CACHE_PATH):
        sys.exit(f"threads_metrics_cache.json が見つかりません: {CACHE_PATH}\n/fetcherを先に実行してください。")

    if not os.path.exists(XLSX_PATH):
        if args.create_if_missing:
            create_workbook(XLSX_PATH)
        else:
            sys.exit(
                f"content_pdca_base.xlsx が見つかりません: {XLSX_PATH}\n"
                "--create-if-missing オプションを付けるか、init_xlsx.py で先に作成してください。"
            )

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    rows = build_rows_from_meta() if args.from_meta else build_rows()
    if not rows:
        src = "post-history-meta.md" if args.from_meta else "threads_metrics_cache.json"
        print(f"書き込むデータがありません（{src}が空または形式不正）。")
        return

    written = write_rows(ws, rows)
    if written == 0:
        print("新規行なし（既存データと重複なし → 差分ゼロ）。")
    else:
        wb.save(XLSX_PATH)
        print(f"{written}行を追加しました → {XLSX_PATH}")


if __name__ == "__main__":
    main()
